from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# مقادیر E24 به صورت رشته برای کنترل کامل
e24_base_str = ["1", "1.1", "1.2", "1.3", "1.5", "1.6", "1.8", "2",
                "2.2", "2.4", "2.7", "3", "3.3", "3.6", "3.9", "4.3",
                "4.7", "5.1", "5.6", "6.2", "6.8", "7.5", "8.2", "9.1"]

scales = [1, 10, 100, 1_000, 10_000, 100_000, 1_000_000]

def format_value(val_str, scale):
    if scale == 1:
        return f"{val_str} Ω"
    elif scale == 1_000:
        return f"{val_str} kΩ"
    elif scale == 1_000_000:
        return f"{val_str} MΩ"
    elif scale == 10:
        num = float(val_str) * 10
        return f"{int(num)} Ω" if num.is_integer() else f"{num:.1f} Ω"
    elif scale == 100:
        num = float(val_str) * 100
        return f"{int(num)} Ω" if num.is_integer() else f"{num:.1f} Ω"
    elif scale == 10_000:
        num = float(val_str) * 10
        return f"{int(num)} kΩ" if num.is_integer() else f"{num:.1f} kΩ"
    elif scale == 100_000:
        num = float(val_str) * 100
        return f"{int(num)} kΩ" if num.is_integer() else f"{num:.1f} kΩ"
    return val_str

# ساخت لیست تمام مقادیر
all_values = ["0 Ω"]
for scale in scales:
    for val in e24_base_str:
        all_values.append(format_value(val, scale))

# تنظیمات اندازه
CM_TO_COL = 5.33
CM_TO_ROW = 37.8

BOX_WIDTH = 4.8 * CM_TO_COL      # 4.8 cm عرض هر جعبه
GAP_WIDTH = 0.5 * CM_TO_COL      # 0.5 cm فاصله افقی
PART_HEIGHT = (2.8 / 3) * CM_TO_ROW  # ارتفاع هر قسمت (2.8 cm ÷ 3)
GAP_HEIGHT = 0.5 * CM_TO_ROW     # 0.5 cm فاصله عمودی

NUM_BOX_COLS = 3  # 3 جعبه در هر ردیف

# سبک border ضخیم یکسان برای همه لبه‌ها
medium = Side(style='medium')

# ایجاد workbook
wb = Workbook()
ws = wb.active

font = Font(name='Arial', size=20, bold=True)
align = Alignment(horizontal="center", vertical="center")

# تنظیم عرض ستون‌ها: 3 جعبه + 2 فاصله
col = 1
for i in range(NUM_BOX_COLS):
    ws.column_dimensions[get_column_letter(col)].width = BOX_WIDTH
    col += 1
    if i < NUM_BOX_COLS - 1:
        ws.column_dimensions[get_column_letter(col)].width = GAP_WIDTH
        col += 1

# محاسبه تعداد کل جعبه‌ها
total_boxes = (len(all_values) + 2) // 3

for box_idx in range(total_boxes):
    # موقعیت ستون جعبه (با فاصله)
    col_pos = (box_idx % NUM_BOX_COLS) * 2 + 1
    # هر بلوک = 3 سطر جعبه + 1 سطر فاصله → ردیف شروع = block * 4 + 1
    block = box_idx // NUM_BOX_COLS
    start_row = block * 4 + 1

    # دریافت سه مقدار جعبه
    vals = []
    for i in range(3):
        idx = box_idx * 3 + i
        vals.append(all_values[idx] if idx < len(all_values) else "")

    # --- سلول بالا ---
    cell_top = ws.cell(start_row, col_pos, vals[0])
    cell_top.font = font
    cell_top.alignment = align
    # border: top, left, right (bottom فقط اگر تک سلولی باشه)
    if vals[1] == "":
        # فقط یک سلول
        cell_top.border = Border(top=medium, bottom=medium, left=medium, right=medium)
    else:
        cell_top.border = Border(top=medium, left=medium, right=medium)

    # --- سلول وسط (اگر وجود داشت) ---
    if vals[1] != "":
        cell_mid = ws.cell(start_row + 1, col_pos, vals[1])
        cell_mid.font = font
        cell_mid.alignment = align
        if vals[2] == "":
            # دو سلول: وسط = پایین
            cell_mid.border = Border(bottom=medium, left=medium, right=medium)
        else:
            # سه سلول: وسط فقط چپ و راست
            cell_mid.border = Border(left=medium, right=medium)

    # --- سلول پایین (اگر وجود داشت) ---
    if vals[2] != "":
        cell_bot = ws.cell(start_row + 2, col_pos, vals[2])
        cell_bot.font = font
        cell_bot.alignment = align
        cell_bot.border = Border(bottom=medium, left=medium, right=medium)

    # تنظیم ارتفاع سطرها
    ws.row_dimensions[start_row].height = PART_HEIGHT
    if vals[1] != "":
        ws.row_dimensions[start_row + 1].height = PART_HEIGHT
    if vals[2] != "":
        ws.row_dimensions[start_row + 2].height = PART_HEIGHT

# تنظیم سطرهای فاصله (هر 4 سطر یکی فاصله)
total_blocks = (total_boxes + NUM_BOX_COLS - 1) // NUM_BOX_COLS
for b in range(total_blocks - 1):  # آخرین بلوک فاصله نمی‌خواد
    gap_row = (b + 1) * 4
    ws.row_dimensions[gap_row].height = GAP_HEIGHT

# تنظیمات چاپ
ws.page_setup.paperSize = 9
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToPage = False
ws.page_setup.scale = 100
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5

# ذخیره
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
filename = f"E24_3Part_Boxes_UniformBorder_{timestamp}.xlsx"
wb.save(filename)
print(f"✅ فایل با جعبه‌های 3 قسمتی و کادر ضخیم یکدست ایجاد شد: {filename}")